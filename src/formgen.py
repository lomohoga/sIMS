from mysql.connector import Error as MySQLError
from datetime import datetime

# to save the file as a stream
from tempfile import NamedTemporaryFile
from os import remove

# to open the form
from openpyxl import load_workbook

# to style text in the form
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from src.blueprints.database import connect_db
from src.blueprints.exceptions import ItemNotFoundError, RequestNotFoundError

def form_58 (db, item):
    db.execute(f"SELECT ItemName, Category, ItemDescription, Unit FROM item WHERE ItemID = '{item}'")
    data = db.fetchone()
    if data is None: raise ItemNotFoundError(item = item)

    #get deliveries
    db.execute(f"SELECT DeliveryDate, DeliveryID, DeliveryQuantity, Time FROM delivery WHERE ItemID = '{item}' ORDER BY DeliveryDate ASC, Time ASC")
    deliveries = db.fetchall()

    #get request items
    db.execute(f"SELECT * FROM (SELECT RequestID, DateReceived, TimeReceived, DeliveryStock, SUM(QuantityIssued), UPPER(RequestedBy), ShelfLife FROM request_item INNER JOIN (SELECT ItemID, ShelfLife, COALESCE(SUM(IF(ShelfLife IS NULL OR DATEDIFF(CURDATE(), ADDDATE(DeliveryDate, ShelfLife)) <= 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) GROUP BY ItemID) AS w USING (ItemID) INNER JOIN request USING (RequestID) WHERE ItemID = '{item}' AND StatusID = 4 GROUP BY ItemID, RequestID ORDER BY RequestID DESC) AS x ORDER BY RequestID ASC")
    requests = db.fetchall()

    wb = load_workbook("./src/form_templates/template_58.xlsx", rich_text = True)
    ws = wb.active

    ws["A8"] = CellRichText(ws["A8"].value, TextBlock(InlineFont(b = True), f"{data[0].upper()} {(data[1]) if data[1] is not None else ''}"))
    ws["A9"] = CellRichText(ws["A9"].value, TextBlock(InlineFont(b = True), data[2]))
    ws["A10"] = CellRichText(ws["A10"].value, TextBlock(InlineFont(b = True), data[3]))
    ws["F8"] = CellRichText(ws["F8"].value, TextBlock(InlineFont(b = True), item))

    i = 0
    j = 0
    ns = None
    continuePage = False
    numRows = 30
    lastDT = ''

    # TODO: Continue numbers in second page
    while(i + j != len(requests) + len(deliveries)):
        lastBalance = 0

        if((i + j) % numRows == 0):
            #Clone worksheet
            if(i + j > 0): continuePage = True
            if(i + j > 0 and continuePage): lastBalance = ns[f"F{13 + numRows - 1}"].value
            ns = wb.copy_worksheet(ws)

        if(i == len(deliveries)):
            ns[f"A{13 + ((i + j) % numRows)}"] = requests[j][1]
            db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE RequestID < {requests[j][0]} && StatusID = 4")
            ris = db.fetchone()[0]
            ns[f"B{13 + ((i + j) % numRows)}"] = f"RIS-{ris}"
            ns[f"D{13 + ((i + j) % numRows)}"] = requests[j][4]
            ns[f"E{13 + ((i + j) % numRows)}"] = requests[j][5]
            curDT = requests[j][1].strftime('%Y-%m-%d') + ' ' + str(requests[j][2])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF(ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0 AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"F{12 + ((i + j) % numRows)}"].value
            ns[f"F{13 + ((i + j) % numRows)}"] = lastBalance - ns[f"D{13 + ((i + j) % numRows)}"].value
            if ((i + j) % numRows) == 0: ns[f"G{13 + ((i + j) % numRows)}"] = requests[j][6]
            j = j + 1
            continue

        if(j == len(requests)):
            ns[f"A{13 + ((i + j) % numRows)}"] = deliveries[i][0]
            ns[f"B{13 + ((i + j) % numRows)}"] = f"D-{deliveries[i][1]}"
            ns[f"C{13 + ((i + j) % numRows)}"] = deliveries[i][2]
            curDT = deliveries[i][0].strftime('%Y-%m-%d') + ' ' + str(deliveries[i][3])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF((ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0) AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"F{12 + ((i + j) % numRows)}"].value
            ns[f"F{13 + ((i + j) % numRows)}"] = lastBalance + ns[f"C{13 + ((i + j) % numRows)}"].value
            i = i + 1
            continue
        
        d = f"{deliveries[i][0]} {deliveries[i][3]}"
        r = f"{requests[j][1]} {requests[j][2]}"

        if d > r:
            ns[f"A{13 + ((i + j) % numRows)}"] = requests[j][1]
            db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE RequestID < {requests[j][0]} && StatusID = 4")
            ris = db.fetchone()[0]
            ns[f"B{13 + ((i + j) % numRows)}"] = f"RIS-{ris}"
            ns[f"D{13 + ((i + j) % numRows)}"] = requests[j][4]
            ns[f"E{13 + ((i + j) % numRows)}"] = requests[j][5]
            curDT = requests[j][1].strftime('%Y-%m-%d') + ' ' + str(requests[j][2])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF(ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0 AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"F{12 + ((i + j) % numRows)}"].value
            ns[f"F{13 + ((i + j) % numRows)}"] = lastBalance - ns[f"D{13 + ((i + j) % numRows)}"].value
            if ((i + j) % numRows) == 0: ns[f"G{13 + ((i + j) % numRows)}"] = requests[j][6]
            j = j + 1
        else:
            ns[f"A{13 + ((i + j) % numRows)}"] = deliveries[i][0]
            ns[f"B{13 + ((i + j) % numRows)}"] = f"D-{deliveries[i][1]}"
            ns[f"C{13 + ((i + j) % numRows)}"] = deliveries[i][2]
            curDT = deliveries[i][0].strftime('%Y-%m-%d') + ' ' + str(deliveries[i][3])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF((ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0) AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"F{12 + ((i + j) % numRows)}"].value
            ns[f"F{13 + ((i + j) % numRows)}"] = lastBalance + ns[f"C{13 + ((i + j) % numRows)}"].value
            i = i + 1

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)
    if(len(requests) + len(deliveries) > 0): wb.remove(ws)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)

def form_59 (db, request):
    db.execute(f"SELECT * FROM request WHERE RequestID = {request}")
    if db.fetchone() is None: raise RequestNotFoundError(request = request)

    db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE StatusID = 4 && hasPropertyApproved = 0 && RequestID < {request};")
    ics_num = str(db.fetchone()[0])

    db.execute(f"SELECT UPPER(RequestedBy) FROM request WHERE RequestID = {request}")
    (req_by,) = db.fetchone()
    db.execute(f"SELECT UPPER(CONCAT(FirstName, ' ', LastName)) FROM request INNER JOIN user ON IssuedBy = Username WHERE RequestID = {request}")
    (iss_by,) = db.fetchone()
    db.execute(f"SELECT DateReceived, DateIssued FROM request WHERE RequestID = {request}")
    (date_rec, date_iss) = db.fetchone()

    db.execute(f"SELECT QuantityIssued, Unit, RequestPrice, RequestPrice * QuantityIssued, ItemDescription, ItemID, ShelfLife FROM request_item INNER JOIN stock USING (ItemID) INNER JOIN request USING (RequestID) WHERE RequestID = '{request}';")
    items = db.fetchall()

    if len(items) == 0: return None

    wb = load_workbook("./src/form_templates/template_59.xlsx", rich_text = True)
    ws = wb.active
    ws["G7"] = CellRichText(ws["G7"].value, TextBlock(InlineFont(b = True), ics_num))

    for i in range(len(items)):
        row = items[i]

        ws[f"A{12 + i}"] = row[0]
        ws[f"B{12 + i}"] = row[1]
        ws[f"C{12 + i}"] = row[2]
        ws[f"D{12 + i}"] = row[3]
        ws[f"E{12 + i}"] = row[4]
        ws[f"G{12 + i}"] = row[5]
        ws[f"H{12 + i}"] = row[6] if row[6] is not None else "\u2014"

    ws["F46"] = req_by
    ws["F50"] = date_rec
    ws["A46"] = iss_by
    ws["A50"] = date_iss

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)

def form_63 (db, request):
    db.execute(f"SELECT * FROM request WHERE RequestID = {request}")
    if db.fetchone() is None: raise RequestNotFoundError(request = request)

    db.execute(f"SELECT RequestID, item.ItemID AS ItemID, item.ItemDescription AS ItemDescription, RequestQuantity, SUM(QuantityIssued), Purpose, Remarks FROM request_item INNER JOIN request USING (RequestID) INNER JOIN item USING (ItemID) INNER JOIN stock USING (ItemID) WHERE RequestID = '{request}' GROUP BY ItemID, RequestID")
    data = db.fetchall()

    db.execute(f"SELECT UPPER(RequestedBy), RequestDate FROM request WHERE RequestID = {request}")
    req_requested = db.fetchone()
    db.execute(f"SELECT UPPER(CONCAT(FirstName, ' ', LastName)) AS IssuedBy, DateIssued FROM request INNER JOIN user ON (Username = IssuedBy) WHERE RequestID = {request}")
    req_issued = db.fetchone()
    db.execute(f"SELECT UPPER(RequestedBy), DateReceived FROM request WHERE RequestID = {request}")
    req_received = db.fetchone()

    db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE RequestID < {request} && StatusID = 4")
    ris = db.fetchone()[0]

    wb = load_workbook("./src/form_templates/template_63.xlsx", rich_text = True)
    ws = wb.active
    ws["G9"] = ris

    for i in range(len(data)):
        req = data[i]

        ws[f"A{12 + i}"] = req[1]
        ws[f"C{12 + i}"] = req[2]
        ws[f"D{12 + i}"] = req[3]
        if req[3] > 0:
            ws[f"E{12 + i}"] = "\u2714"
        else: ws[f"F{12 + i}"] = "\u2714"
        ws[f"G{12 + i}"] = req[4]
        if req[6] is not None:
            ws[f"H{12 + i}"] = req[6]
        ws["B32"] = req[5]

    ws["C37"] = req_requested[0]
    ws["C39"] = req_requested[1]
    ws["D37"] = "ROSA TAYAMORA"
    ws["F37"] = req_issued[0]
    ws["F39"] = req_issued[1]
    ws["H37"] = req_received[0]
    ws["H39"] = req_received[1]

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)

def form_69 (db, item):
    db.execute(f"SELECT ItemName, Category, ItemDescription FROM item WHERE ItemID = '{item}'")
    data = db.fetchone()
    if data is None: raise ItemNotFoundError(item = item)

    #get deliveries
    db.execute(f"SELECT DeliveryDate, DeliveryID, DeliveryQuantity, Time, DeliveryQuantity * DeliveryPrice FROM delivery WHERE ItemID = '{item}' ORDER BY DeliveryID ASC")
    deliveries = db.fetchall()

    #get request items
    db.execute(f"SELECT * FROM (SELECT RequestID, DateReceived, TimeReceived, DeliveryStock, QuantityIssued, UPPER(RequestedBy), RequestPrice * QuantityIssued, Remarks FROM request_item INNER JOIN (SELECT ItemID, Price, COALESCE(SUM(IF(ShelfLife IS NULL OR DATEDIFF(CURDATE(), ADDDATE(DeliveryDate, ShelfLife)) <= 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) GROUP BY ItemID) AS w USING (ItemID) INNER JOIN request USING (RequestID) WHERE ItemID = '{item}' AND StatusID = 4 ORDER BY RequestID DESC) AS x ORDER BY RequestID ASC")
    requests = db.fetchall()

    wb = load_workbook("./src/form_templates/template_69.xlsx", rich_text = True)
    ws = wb.active

    ws["B8"] = CellRichText(ws["B8"].value, TextBlock(InlineFont(b = True), f"{data[0].upper()} {(data[1]) if data[1] is not None else ''}"))
    ws["B10"] = CellRichText(ws["B10"].value, TextBlock(InlineFont(b = True), data[2]))
    ws["I9"] = CellRichText(ws["I9"].value, TextBlock(InlineFont(b = True), item))

    i = 0
    j = 0
    ns = None
    continuePage = False
    numRows = 18
    lastDT = ''

    while(i + j != len(requests) + len(deliveries)):
        lastBalance = 0

        if((i + j) % numRows == 0):
            #Clone worksheet
            if(i + j > 0): continuePage = True
            if(i + j > 0 and continuePage): lastBalance = ns[f"F{13 + numRows - 1}"].value
            ns = wb.copy_worksheet(ws)

        if(i == len(deliveries)):
            ns[f"B{13 + ((i + j) % numRows)}"] = requests[j][1]
            db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE hasPropertyApproved = 1 && RequestID < {requests[j][0]}")
            ns[f"C{13 + ((i + j) % numRows)}"] = f"PAR-{db.fetchone()[0]}"
            ns[f"E{13 + ((i + j) % numRows)}"] = requests[j][4]
            ns[f"F{13 + ((i + j) % numRows)}"] = requests[j][5]
            curDT = requests[j][1].strftime('%Y-%m-%d') + ' ' + str(requests[j][2])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF(ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0 AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"H{12 + ((i + j) % numRows)}"].value
            ns[f"H{13 + ((i + j) % numRows)}"] = lastBalance - ns[f"E{13 + ((i + j) % numRows)}"].value
            ns[f"I{13 + ((i + j) % numRows)}"] = requests[j][6]
            ns[f"J{13 + ((i + j) % numRows)}"] = requests[j][7]
            j = j + 1
            continue

        if(j == len(requests)):
            ns[f"B{13 + ((i + j) % numRows)}"] = deliveries[i][0]
            ns[f"C{13 + ((i + j) % numRows)}"] = f"D-{deliveries[i][1]}"
            ns[f"D{13 + ((i + j) % numRows)}"] = deliveries[i][2]
            curDT = deliveries[i][0].strftime('%Y-%m-%d') + ' ' + str(deliveries[i][3])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF((ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0) AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"H{12 + ((i + j) % numRows)}"].value
            ns[f"H{13 + ((i + j) % numRows)}"] = lastBalance + ns[f"D{13 + ((i + j) % numRows)}"].value
            ns[f"I{13 + ((i + j) % numRows)}"] = deliveries[i][4]
            i = i + 1
            continue
        
        d = f"{deliveries[i][0]} {deliveries[i][3]}"
        r = f"{requests[j][1]} {requests[j][2]}"

        if d > r:
            ns[f"B{13 + ((i + j) % numRows)}"] = requests[j][1]
            db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE hasPropertyApproved = 1 && RequestID < {requests[j][0]}")
            ns[f"C{13 + ((i + j) % numRows)}"] = f"PAR-{db.fetchone()[0]}"
            ns[f"E{13 + ((i + j) % numRows)}"] = requests[j][4]
            ns[f"F{13 + ((i + j) % numRows)}"] = requests[j][5]
            curDT = requests[j][1].strftime('%Y-%m-%d') + ' ' + str(requests[j][2])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF(ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0 AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"H{12 + ((i + j) % numRows)}"].value
            ns[f"H{13 + ((i + j) % numRows)}"] = lastBalance - ns[f"E{13 + ((i + j) % numRows)}"].value
            ns[f"I{13 + ((i + j) % numRows)}"] = requests[j][6]
            ns[f"J{13 + ((i + j) % numRows)}"] = requests[j][7]
            j = j + 1
        else:
            ns[f"B{13 + ((i + j) % numRows)}"] = deliveries[i][0]
            ns[f"C{13 + ((i + j) % numRows)}"] = f"D-{deliveries[i][1]}"
            ns[f"D{13 + ((i + j) % numRows)}"] = deliveries[i][2]
            curDT = deliveries[i][0].strftime('%Y-%m-%d') + ' ' + str(deliveries[i][3])
            if curDT != lastDT:
                db.execute(f"SELECT (SELECT COALESCE(SUM(IF((ShelfLife IS NULL OR TIMESTAMPDIFF(SECOND, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime), '{curDT}') > 0) AND TIMESTAMPDIFF(SECOND, TIMESTAMPADD(DAY, ShelfLife, CAST(CONCAT(DeliveryDate, ' ', Time) AS datetime)), '{curDT}') < 0, DeliveryQuantity, 0)), 0) AS DeliveryStock FROM item LEFT JOIN delivery USING (ItemID) WHERE ItemID = '{item}') - (SELECT COALESCE(SUM(IF(TIMESTAMPDIFF(SECOND, CAST(CONCAT(DateReceived, ' ', TimeReceived) AS datetime), '{curDT}') > 0, QuantityIssued, 0)), 0) FROM request_item LEFT JOIN request USING (RequestID) WHERE StatusID = 4 AND ItemID = '{item}') AS Stock")
                lastBalance = db.fetchone()[0]
                lastDT = curDT
            else:
                if (i + j) % numRows == 0 and continuePage: pass
                else: lastBalance = ns[f"H{12 + ((i + j) % numRows)}"].value
            ns[f"H{13 + ((i + j) % numRows)}"] = lastBalance + ns[f"D{13 + ((i + j) % numRows)}"].value
            ns[f"I{13 + ((i + j) % numRows)}"] = deliveries[i][4]
            i = i + 1

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)
    if(len(requests) + len(deliveries) > 0): wb.remove(ws)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)

def form_71 (db, request):
    db.execute(f"SELECT * FROM request WHERE RequestID = {request}")
    if db.fetchone() is None: raise RequestNotFoundError(request = request)

    db.execute(f"SELECT COUNT(*) + 1 FROM request WHERE StatusID = 4 && hasPropertyApproved = 1 && RequestID < {request};")
    par_num = str(db.fetchone()[0])

    db.execute(f"SELECT UPPER(RequestedBy) FROM request WHERE RequestID = {request}")
    (req_by,) = db.fetchone()
    db.execute(f"SELECT UPPER(CONCAT(FirstName, ' ', LastName)) FROM request INNER JOIN user ON IssuedBy = Username WHERE RequestID = {request}")
    (iss_by,) = db.fetchone()
    db.execute(f"SELECT DateReceived, DateIssued FROM request WHERE RequestID = {request}")
    (date_rec, date_iss) = db.fetchone()

    db.execute(f"SELECT QuantityIssued, Unit, ItemDescription, ItemID, RequestPrice * QuantityIssued FROM request_item INNER JOIN stock USING (ItemID) INNER JOIN request USING (RequestID) WHERE RequestID = '{request}';")
    items = db.fetchall()

    if len(items) == 0: return None

    wb = load_workbook("./src/form_templates/template_71.xlsx", rich_text = True)
    ws = wb.active
    ws["E7"] = CellRichText(ws["E7"].value, TextBlock(InlineFont(b = True), par_num))

    for i in range(len(items)):
        row = items[i]

        ws[f"A{11 + i}"] = row[0]
        ws[f"B{11 + i}"] = row[1]
        ws[f"C{11 + i}"] = row[2]
        ws[f"D{11 + i}"] = row[3]
        ws[f"E{11 + i}"] = date_rec
        ws[f"F{11 + i}"] = row[4]

    ws["A45"] = req_by
    ws["A49"] = date_rec
    ws["D45"] = iss_by
    ws["D49"] = date_iss

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)

def form_73 (items):
    wb = load_workbook("./src/form_templates/template_73.xlsx", rich_text = True)
    ws = wb.active

    ws["G6"] = datetime.today().strftime("%d/%m/%y %H:%M:%S")

    ns = None
    numRows = 20
    
    for i in range(len(items)):
        if(i % numRows == 0): ns = wb.copy_worksheet(ws)

        ns[f"C{15 + (i % numRows)}"] = items[i]["ItemName"]
        ns[f"D{15 + (i % numRows)}"] = items[i]["ItemDescription"]
        ns[f"E{15 + (i % numRows)}"] = items[i]["ItemID"]
        ns[f"F{15 + (i % numRows)}"] = items[i]["Unit"]
        ns[f"G{15 + (i % numRows)}"] = items[i]["Price"]
        ns[f"H{15 + (i % numRows)}"] = items[i]["AvailableStock"]

    file = NamedTemporaryFile(suffix = ".xlsx", delete = False)
    if(len(items) > 0): wb.remove(ws)

    try:
        with file as f:
            wb.save(f.name)
            f.seek(0)
            return f.read()
    finally:
        remove(file.name)